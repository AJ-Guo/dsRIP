import re
from openpyxl import Workbook
import csv
import gc
import os
from statistics import mean
import sqlite3
import ORF_finder_module
import numpy as np
import RNA
import RNAtweaks.RNAtweaks as RNAtweaks
from Bio import Align
from Bio import SeqIO
from Bio import SeqIO
from Bio.Seq import Seq
from orffinder import orffinder
from Bio.SeqRecord import SeqRecord
import pandas as pd
import matplotlib
from indexer_for_off_target_bowtie1 import run_cd_hit_and_indexer
import matplotlib.pyplot as plt
import numpy as np
from off_target_prediction_siRNA import off_target_siRNA_all
import itertools
import heapq
import itertools
import heapq
import pandas as pd
import re
import matplotlib.pyplot as plt
from ORF_finder_module import correct_sequence_based_on_ORF

matplotlib.use('Agg')
def sanitize_sheet_title(title):
    invalid_chars = [':', '*', '?', '/', '\\', '[', ']']
    for char in invalid_chars:
        title = title.replace(char, '_')  # replace invalid characters with underscore
    return title

def best_region_all(fasta, off_target_species, user_dir, min_length=300, max_length=800, user_siRNA_length=21,
                    user_ORF_corr=False, save_off_targets_plot=True, only_ORF=True, bufsi=21, safety=True,
                    safety_essential_coeff=20, efficacy_coeff=50, safety_coeff=50, new_off_target_on = False, user_mismatch = 0, new_off_target_species_input_name="user_transcriptome"):


    if new_off_target_on:
        run_cd_hit_and_indexer(user_dir + "/efficiency", c_param=0.95, min_length=200, user_species_name=new_off_target_species_input_name)
    else:
        pass


    min_length = min_length + 1
    max_length = max_length + 1
    final_clean = {}
    siRNA_length = int(user_siRNA_length)

    min_window = int(min_length)
    max_window = int(max_length)

    file_path = fasta

    try:
        sequence_data = read_fasta_file(file_path)
    except Exception as e:
        print(f"An error occurred while reading the FASTA file: {e}")
        return

    fasta_dict = {}
    fasta_dict_score = {}

    gene_count = 0  # Initialize a counter for processed genes

    for name, sequence in sequence_data:
        try:
            # Limit processing to the first 5 genes
            if gene_count >= 3:
                break

            # Check for valid DNA sequence.
            if not check_valid_dna(sequence):
                print("\nInvalid DNA sequence detected:", sequence)
                continue

            name = sanitize_sheet_title(name)

            if user_ORF_corr:
                corrected_sequence = correct_sequence_based_on_ORF(sequence, name)
                gene_ORF = ORF_finder_module.get_ORF(corrected_sequence, name)
                gene_ORF = str(gene_ORF[4])

            else:
                corrected_sequence = sequence
                gene_ORF = ORF_finder_module.get_ORF(corrected_sequence, name)
                gene_ORF = str(gene_ORF[4])

            cor_seq_RNA = corrected_sequence.replace("T", "U")

            seq_length = len(corrected_sequence)

            ORF_info = ORF_siRNA(cor_seq_RNA)

            siRNAs, siRNAs_DNA_off_target = generate_siRNA(cor_seq_RNA, siRNA_length=siRNA_length)

            if safety:
                try:
                    per_species_summary, all_off_targets_sum, lethal_off_targets_sum, all_siRNAs_report, off_targets_plot = off_target_siRNA_all(
                        siRNAs_DNA_off_target, seq_length, name, off_target_species, user_dir, siRNA_length= user_siRNA_length, user_mismatch=user_mismatch,new_off_target_species_input_name=new_off_target_species_input_name)

                    if save_off_targets_plot:
                        plt.savefig(user_dir + "/efficiency/off_target/" + name + '_off_target_plot.png')
                except Exception as e:
                    print(f"An error occurred during off-target analysis for {name}: {e}")
                    all_siRNAs_report = []
            else:
                all_siRNAs_report = []

            access = mRNA_accessibility(cor_seq_RNA, siRNA_length=siRNA_length)

            gene_siRNA_feature = siRNA_feature_prediction(siRNAs, cor_seq_RNA, name, siRNA_length, access, ORF_info,
                                                          all_siRNAs_report)

            gene_dict = {
                "gene_name": name,
                "input_gene_seq": str(sequence).upper(),
                "corrected_gene_seq": str(corrected_sequence),
                "gene_ORF": str(gene_ORF),
                "gene_RNA_seq": str(cor_seq_RNA),
                "siRNA_features": gene_siRNA_feature
            }

            scored_features = score_and_update_siRNA_features(gene_dict, safety=safety,
                                                              safety_essential_coeff=safety_essential_coeff)

            best_region = find_best_siRNA_region(scored_features, min_window=(min_window - (siRNA_length - 1)),
                                                 max_window=(max_window - (siRNA_length - 1)), only_ORF=only_ORF,
                                                 safety=safety, efficiency_priority=efficacy_coeff, safety_priority=safety_coeff, user_dir=user_dir)
            #print(best_region)

            gene_dict_score = {
                "gene_name": name,
                "input_gene_seq": str(sequence).upper(),
                "corrected_gene_seq": str(corrected_sequence),
                "gene_ORF": str(gene_ORF),
                "gene_RNA_seq": str(cor_seq_RNA),
                "siRNA_features": scored_features
            }

            fasta_dict[name] = gene_dict
            fasta_dict_score[name] = gene_dict
            best_dsRNA_region = extract_subsequence(str(corrected_sequence), best_region, bufsi=bufsi)

            final_clean[name] = {
                "gene_name": name,
                "corrected_gene_seq": str(corrected_sequence),
                "gene_ORF": str(gene_ORF),
                "optimized_dsRNA": best_dsRNA_region
            }

            gene_count += 1  # Increment the counter

        except Exception as e:
            print(f"An error occurred while processing gene {name}: {e}")
            continue

    # Open the file in write mode and save the content
    #try:
    #    dict_to_excel(final_clean, user_dir + "/efficiency/dsRNA_designs_summary.xlsx")
   # except Exception as e:
    #    print(f"An error occurred while saving dsRNA designs summary: {e}")

    try:
        write_fasta_dict_to_excel(fasta_dict, user_dir)
    except Exception as e:
        print(f"An error occurred while saving FASTA dictionary to Excel: {e}")

    ###FASTA OUTPUT###
    fasta_output = ""

    for name, data in final_clean.items():
        # Append the FASTA entry for each gene to the fasta_output string
        fasta_output += f">{data['gene_name']} \n{data['optimized_dsRNA']}\n"

    # Specify the output file name for the FASTA data
    fasta_filename = user_dir + "/efficiency/best_dsRNA_regions.fasta"

    # Write the accumulated FASTA formatted string to a file
    try:
        with open(fasta_filename, 'w') as file:
            file.write(fasta_output)
    except Exception as e:
        print(f"An error occurred while saving the FASTA output: {e}")


def dict_to_excel(data_dict, output_filename):
    """Converts a dictionary of gene data to an Excel file.

    Args:
    - data_dict (dict): Dictionary containing gene data.
    - output_filename (str): Desired name for the output Excel file.

    Returns:
    None. Writes the data to an Excel file.
    """

    # Convert dictionary to pandas DataFrame
    df = pd.DataFrame.from_dict(data_dict, orient='index')

    # Write DataFrame to Excel
    df.to_excel(output_filename, engine='openpyxl')





def write_fasta_dict_to_excel(fasta_dict,user_dir,excel_siRNA_predictions=True):

    if excel_siRNA_predictions:
    # Initialize a new Excel workbook
        wb = Workbook()
        # Removing the default created sheet
        wb.remove(wb.active)



        # Iterate through the fasta_dict
        for gene_name, gene_dict in fasta_dict.items():
            # Create a new sheet for the gene_name
            cleaned_gene_name = ''.join([char if char.isalnum() or char == ' ' else '_' for char in gene_name])
            #primary_sheet = wb.create_sheet(cleaned_gene_name)


            primary_data = ['gene_name', 'input_gene_seq', 'corrected_gene_seq', 'gene_ORF', 'gene_RNA_seq']
            #for idx, item in enumerate(primary_data, start=1):
            #    primary_sheet.cell(row=idx, column=1, value=item)
            #    primary_sheet.cell(row=idx, column=2, value=gene_dict.get(item, ''))

            # Create a new sheet for siRNA features
            siRNA_sheet = wb.create_sheet(cleaned_gene_name[:10]+cleaned_gene_name[-5:])
            siRNA_features = gene_dict.get('siRNA_features', {})

            # First, let's set headers based on the first dictionary inside 'siRNA_features'
            first_item = next(iter(siRNA_features.values()), {})
            if isinstance(first_item, dict):
                for sub_idx, header in enumerate(first_item.keys(), start=2):
                    siRNA_sheet.cell(row=1, column=sub_idx, value=header)

            # Write siRNA data
            for idx, (key, value) in enumerate(siRNA_features.items(), start=2):
                siRNA_sheet.cell(row=idx, column=1, value=key)
                if isinstance(value, dict):
                    # If the value is also a dictionary, write its content in the following columns
                    for sub_idx, (sub_key, sub_value) in enumerate(value.items(), start=2):
                        siRNA_sheet.cell(row=idx, column=sub_idx, value=str(sub_value))
                else:
                    siRNA_sheet.cell(row=idx, column=2, value=str(value))

        # Set the first sheet as the active one
        if wb.worksheets:  # Check if there's at least one sheet
            wb.active = 0

        # Save the workbook
        wb.save(user_dir +"/efficiency/siRNA_predictions.xlsx")






def get_value(key,cursor):

    cursor.execute("SELECT value FROM lookup WHERE key=?", (key,))
    result = cursor.fetchone()


    # Return the value if the key was found, otherwise return None
    return result[0] if result else None


def reverseComp_RNA(seq):
    seq_to_revCom = Seq (seq)
    revCom = seq_to_revCom.reverse_complement_rna()

    return str(revCom)


def Comp_RNA(seq):
    seq_to_revCom = Seq (seq)
    Com = seq_to_revCom.complement_rna()

    return str(Com)

def get_dict_with_highest_length(list_of_dicts, gene):
    if not list_of_dicts:
        return None

    dict_with_highest_length = max(list_of_dicts, key=lambda d: d.get('length', 0))

    if dict_with_highest_length["sense"] == "+":
        dict_with_highest_length["gene_sense_direction"] = gene.seq
        dict_with_highest_length["id"] = gene.id
        return dict_with_highest_length
    elif dict_with_highest_length["sense"] == "-":
        gene_seq_sense = Seq(gene.seq).reverse_complement()
        dict_with_highest_length["gene_sense_direction"] = gene_seq_sense
        dict_with_highest_length["id"] = gene.id
        return dict_with_highest_length


def read_fasta_file(file_path):
    """
    Read a working_fasta file, ensure they are valid DNA sequences and return them in uppercase.
    """
    try:
        with open(file_path, "r") as handle:
            sequence_data = [(record.id, record.seq.upper()) for record in SeqIO.parse(handle, "fasta")]
    except Exception as e:
        print(f"Error reading the FASTA file: {e}")
        sequence_data = []

    return sequence_data



def check_valid_dna(seq: str) -> bool:
    """
    Check if the input sequence is a valid DNA sequence.
    """
    valid_bases = set("ATUCG")
    return set(seq) <= valid_bases


def dna_to_rna(dna_seq: str) -> str:
    """
    Convert a DNA sequence to an RNA sequence.
    """
    return dna_seq.replace('T', 'U')







def kmer_pairs_np(string1, string2, k):

    siRNA_length = k-2
    assert len(string1) == len(string2), "Both strings must be of the same length."

    arr1 = np.array(list(string1))
    arr2 = np.array(list(string2))

    num_kmers = len(string1) - k + 1
    kmers1 = np.lib.stride_tricks.sliding_window_view(arr1, window_shape=k)
    kmers2 = np.lib.stride_tricks.sliding_window_view(arr2, window_shape=k)

    combined = list(zip(map(''.join, kmers1), map(''.join, kmers2)))
    overhanged_siRNAs = [(tup[0][2:], tup[1][:-2]) for tup in combined]

    #print(overhanged_siRNAs)

    return overhanged_siRNAs

def generate_siRNA(sequence, siRNA_length):
    whole_antisense_5_3 = reverseComp_RNA(sequence)

    whole_antisense_5_3_XX=  "XX"+whole_antisense_5_3
    whole_sense_XX_3_5 = "XX"+Comp_RNA(whole_antisense_5_3)

    #print(whole_antisense_5_3_XX,whole_sense_XX_3_5)

    siRNA_tuples = kmer_pairs_np(whole_antisense_5_3_XX,whole_sense_XX_3_5, (siRNA_length+2))


    siRNA_tuples_corr = siRNA_tuples.reverse()



    siRNA_tuples_DNA = [(pair[0].replace("U", "T"), pair[1].replace("U", "T")) for pair in siRNA_tuples]

    siRNAs_for_off_target = [pair[0] for pair in siRNA_tuples_DNA]

    #print(siRNAs_for_off_target)

    return siRNA_tuples, siRNAs_for_off_target

def count_nucleotides(sequence):
    return {
        'A': sequence.count('A'),
        'U': sequence.count('U'),
        'G': sequence.count('G'),
        'C': sequence.count('C')
    }

def thermoAsym(siRNA_tuple,cursor):
    antisense_5_3=siRNA_tuple[0]
    sense_3_5=siRNA_tuple[1]

    try:
        antisense_value = get_value(antisense_5_3[:4],cursor)

        sense_value = get_value((sense_3_5[-4:])[::-1],cursor)
        asymScore = antisense_value - sense_value
        #print(antisense_5_3[:4],(sense_3_5[-4:])[::-1],antisense_5_3, asymScore)
        return asymScore



    except Exception as e:  # Catch all exceptions but it's better to specify the type of exception if known
        print(f"{siRNA_tuple} encountered an error: {e}")

        return "NA"


### Predictions



def mRNA_accessibility(sequence, siRNA_length):
    window = 80  # winsize option of RNAplfold
    span = 40  # span option of RNAplfold
    region = 16  # ulength option of RNAplfold

    sequence = sequence
    api_result = RNAtweaks.api_rnaplfold(sequence, window, span, region=region, temperature=25)

    array = api_result.numpy_array

    df = pd.DataFrame(array)

    #print(df)



    # Assuming the last column of df contains the accessibility data
    accessibility_column = df.iloc[:, -1]

    # Compute the number of 16-nt windows within the given siRNA_length
    num_windows = siRNA_length - 15

    # Sliding over the sequence in siRNA_length windows and computing average accessibility
    avg_accessibilities = []

    for i in range(len(sequence) - siRNA_length + 1):
        siRNA_accessibilities = []
        for j in range(num_windows):
            start = i + j
            end = start + 16
            siRNA_accessibilities.append(accessibility_column[start:end].mean())

        avg_accessibility = np.mean(siRNA_accessibilities)
        avg_accessibilities.append(avg_accessibility)
        #(avg_accessibility)


    #avg_accessibilities = avg_accessibilities[2:]

    #print((avg_accessibilities))
    return avg_accessibilities


def ORF_siRNA(mRNA_seq):
    ORF_info = ORF_finder_module.process_sequence(str(mRNA_seq).replace("U", "T"), "gene")
    return ORF_info



def find_substring(s):
    substrings = ["GGGG", "CCCC"]
    for sub in substrings:
        if sub in s:
            return sub  # Return the found substring
    return None  # Return None if no substring was found

def siRNA_feature_prediction(siRNAs,mRNA_seq,mRNA_name,siRNA_length,access,ORF_once,off_targets=[],temp=25):

    db_path = 'constant_files/lookup.db'
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    md = RNA.md()
    md.temperature = temp

    siRNA_counter = 0

    gene_siRNA_feature = {}

    for siRNA_duplex in siRNAs:



        antisense = siRNA_duplex[0]

        antisense_DNA = str(antisense).replace("U","T")
        sense= siRNA_duplex[1]



        #sense_5_3_DNA = str(sense[::-1]).replace("U","T")

        antisense_9_14 = antisense[8:14]


        antisense_9_14_C = count_nucleotides(antisense_9_14)["C"]
        antisense_9_14_G = count_nucleotides(antisense_9_14)["G"]
        antisense_9_14_GC_perc = round(((antisense_9_14_C + antisense_9_14_G) / 6 * 100),1)

        antisense_total_GC = round(((count_nucleotides(antisense)["C"] + count_nucleotides(antisense)["G"]) / len(antisense) * 100),1)

        ####thermoScore


        asymScore = round(thermoAsym(siRNA_duplex,cursor),3)

        ### RNA fold

        fc = RNA.fold_compound(str(antisense), md)
        (fold_structure, fold_energy) = fc.mfe()




        siRNA_access = access[siRNA_counter]

        code_n = siRNA_counter+1

        ORF_info = ORF_once

        if_ORF = "partial_ORF"

        try:
            if int(ORF_info[1]) < code_n and int(ORF_info[2]) > (siRNA_length + code_n - 1):
                if_ORF = "ORF"

            elif int(ORF_info[1]) > code_n and int(ORF_info[1]) > (siRNA_length + code_n - 1):
                if_ORF = "5_UTR"

            elif int(ORF_info[2]) < code_n and int(ORF_info[2]) < (siRNA_length + code_n - 1):
                if_ORF = "3_UTR"

        except Exception as e:
            if_ORF = "ORF"

        anti_10th_n_A = "0"

        if antisense[9] == "A":
            anti_10th_n_A = "1"


        edge_asym = 0

        if antisense[0] == "A" or antisense[0] == "U" :
            edge_asym += 1

        if antisense[18] == "G" or antisense[0] == "C" :
            edge_asym += 1


        bad_substring = None

        found_substring = find_substring(antisense)

        if found_substring:
            bad_substring = found_substring

        indv_siRNA_dict = {"antisense_5_3":antisense, "sense_3_5_":sense,"antisense_5_3_DNA":antisense_DNA,"anti_GC":antisense_total_GC,"anti_9_14_GC":antisense_9_14_GC_perc,"asymScore":asymScore, "self_fold_energy":round(fold_energy,4),"accessibility":siRNA_access,"anti_10th_A": anti_10th_n_A,"ORF_info":if_ORF ,"G_C_repeat":bad_substring,"edge_asyym":edge_asym,"self_fold_structure":fold_structure}

        siRNA_code = (str(mRNA_name) + "_" + str(siRNA_counter+1) + "_" + str(siRNA_counter + siRNA_length-1))
        siRNA_counter += 1

        gene_siRNA_feature[siRNA_code] =indv_siRNA_dict

        #print(siRNA_code,RNA.co_pf_fold(str(antisense)+"&"+str(mRNA_seq)))

        #"sense_5_3_DNA":sense_5_3_DNA,

    #print(gene_siRNA_feature, off_targets)

    # Check if the lengths match
    if len(gene_siRNA_feature) == len(off_targets):
        # Convert the gene_siRNA_feature keys into a list to index them by order
        siRNA_keys = list(gene_siRNA_feature.keys())

        # Iterate through off_targets by index
        for i, (siRNA_name, all_off_targets, lethal_off_targets) in enumerate(off_targets):
            # Using the index, access the corresponding siRNA key in gene_siRNA_feature
            gene_key = siRNA_keys[i]

            # Update the dictionary for the siRNA in gene_siRNA_feature
            gene_siRNA_feature[gene_key]['all_off_targets'] = all_off_targets
            gene_siRNA_feature[gene_key]['lethal_off_targets'] = lethal_off_targets
    else:
        print("The number of siRNAs in the list and tuples does not match.")
        print(f"gene_siRNA_feature: {len(gene_siRNA_feature)}, off_targets: {len(off_targets)}")

    #print(gene_siRNA_feature)

    return gene_siRNA_feature



def read_parameters(file_path="constant_files/siRNA_parameters.txt"):
    params = {}
    with open(file_path, 'r') as file:
        for line in file:
            key, value = line.strip().split('=')
            params[key] = float(value)
    return params


def normalize_score(value, min_val, max_val, target_min, target_max):
    """
    Normalize a value from its original range (min_val, max_val) to a new target range (target_min, target_max).
    """
    return (value - min_val) / (max_val - min_val) * (target_max - target_min) + target_min

def score_and_update_siRNA_features(input_dict,safety, safety_essential_coeff):
    params = read_parameters()
    asym_para = params['asymScore_weight']
    self_fold_para = params['self_fold_energy_weight']
    #access_para = params['mRNA_accessibility_weight']
    asym_self_access_comb = asym_para + self_fold_para
    GG_CC_para = params['GGGG_CCCC_penalty']
    edge_bonus = params['edge_asyym_bonus']
    anti_GC_bonus = params['anti_GC_bonus']
    anti_9_14_GC_bonus = params['anti_9_14_GC_bonus']
    anti_GC_upper_bound = params['anti_GC_upper_bound']
    anti_9_14_GC_lower_bound = params['anti_9_14_GC_lower_bound']
    ORF_bonus = params['ORF_info_bonus']
    anti_10th_A_bonus= params['anti_10th_A_bonus']

    siRNA_features = input_dict['siRNA_features']

    off_target_scores = []


    for _, features in siRNA_features.items():
        features['asymScore_normalized'] = normalize_score(features['asymScore'], 7.5, -7.5, 100, 0)

        features['self_fold_energy_normalized'] = normalize_score(features['self_fold_energy'], 0, -15, 100, 0)


        # Normalize accessibility: use logarithmic scaling due to wide range (10^-12 -> 0, 1 -> 100)
        if features['accessibility'] > 0:
            log_access = np.log10(features['accessibility'])
            log_min = np.log10(1e-10)
            log_max = np.log10(1)
            features['accessibility_normalized'] = normalize_score(log_access, log_min, log_max, 0, 100)
        else:
            features['accessibility_normalized'] = 0

        # Score from normalized values
        base_score = ((features['asymScore_normalized'] * asym_para) +
                      (features['self_fold_energy_normalized'] * self_fold_para) ) / (asym_self_access_comb)

        #+ (features['accessibility_normalized'] * access_para)

        if safety:
            try:
                all_off_targets_pre = features.get("all_off_targets", 0)
                lethal_off_targets_pre = features.get("lethal_off_targets", 0)
                all_off_targets = all_off_targets_pre - lethal_off_targets_pre
                lethal_off_targets = lethal_off_targets_pre * int(safety_essential_coeff)

                off_target_both_pre_score = all_off_targets + lethal_off_targets
                off_target_scores.append(off_target_both_pre_score)

            except:
                off_target_both_pre_score = 0
                off_target_scores.append(off_target_both_pre_score)
                print("Cannot get off-target scores")

        # Initialize additional score
        additional_score = 0

        if features.get("GGGG_CCCC", None) == "GGGG" or features.get("GGGG_CCCC", None) == "CCCC":
            additional_score -= GG_CC_para


        if features.get("edge_asyym"):
            additional_score += edge_bonus

        # Check conditions to add points
        #if params['anti_GC_lower_bound'] <= features.get("anti_GC", 0) <= params['anti_GC_upper_bound']:
        #    additional_score += params['anti_GC_bonus']

        if features.get("anti_GC", 0) <= anti_GC_upper_bound:
            additional_score += anti_GC_bonus


        if anti_9_14_GC_lower_bound <= features.get("anti_9_14_GC", 0):
            additional_score += anti_9_14_GC_bonus


        if features.get("ORF_info") == "ORF":
            additional_score += ORF_bonus

        if int(features.get("anti_10th_A", 0)) == 1:
            additional_score += anti_10th_A_bonus


        total_score = base_score + additional_score
        features['total_score'] = round(total_score, 1)

    if off_target_scores:
        min_off_target = min(off_target_scores)
        max_off_target = max(off_target_scores)
        range_off_target = max_off_target - min_off_target

        for _, features in siRNA_features.items():
            if safety:
                all_off_targets_pre = features.get("all_off_targets", 0)
                lethal_off_targets_pre = features.get("lethal_off_targets", 0)
                all_off_targets = all_off_targets_pre - lethal_off_targets_pre
                lethal_off_targets = lethal_off_targets_pre * int(safety_essential_coeff)
                off_target_both_pre_score = all_off_targets + lethal_off_targets

                if range_off_target > 0:
                    # Normalize such that min_off_target becomes 100 and max_off_target becomes 0
                    norm_off_target = (1 - (off_target_both_pre_score - min_off_target) / range_off_target) * 100
                else:
                    norm_off_target = 100  # If all scores are identical, set to 100

                features['norm_off_target'] = round(norm_off_target, 1)


    return {'siRNA_features': siRNA_features}




def find_best_siRNA_region(siRNA_scores, min_window, max_window, only_ORF, safety, safety_priority, efficiency_priority, user_dir):
    best_avg_score = float('-inf')
    best_start = 0
    best_end = 0
    best_window_size = 0
    window_data = []  # List to store information about each window



    if only_ORF:
        siRNA_names = [name for name in siRNA_scores['siRNA_features'] if
                       siRNA_scores['siRNA_features'][name]['ORF_info'] == 'ORF']
    else:
        siRNA_names = list(siRNA_scores['siRNA_features'].keys())

    # Exit early if no siRNA names are available
    if not siRNA_names:
        return {
            'start_siRNA': None,


            'end_siRNA': None,
            'average_score': None,
            'window_size': 0,
            "gene_average": None
        }

    # Calculate combined scores based on the weighted total_score and norm_off_target
    siRNA_combined_scores = {}
    total_scores = {}
    norm_off_target_scores = {}
    #print(siRNA_scores)['siRNA_features']
    for name in siRNA_names:
        total_score = siRNA_scores['siRNA_features'][name]['total_score']
        norm_off_target = siRNA_scores['siRNA_features'][name].get('norm_off_target', 0)

        # Compute the combined score using safety_priority and efficiency_priority

        if safety:

            combined_score = (int(efficiency_priority) * int(total_score) + int(safety_priority) * int(norm_off_target)) / 100
            siRNA_combined_scores[name] = combined_score
            total_scores[name] = total_score
            norm_off_target_scores[name] = norm_off_target
        elif not safety:

            combined_score = total_score
            siRNA_combined_scores[name] = combined_score
            total_scores[name] = total_score
            norm_off_target_scores[name] = norm_off_target


    gene_average = sum(siRNA_combined_scores.values()) / len(siRNA_names)

    # Sliding window to find the best region based on the combined scores
    for window_size in range(min_window, max_window + 1):
        current_sum = sum([siRNA_combined_scores[name] for name in siRNA_names[:window_size]])
        for start in range(len(siRNA_names) - window_size + 1):
            end = start + window_size
            current_avg_score = current_sum / window_size

            window_data.append({
                'start_siRNA': siRNA_names[start],
                'end_siRNA': siRNA_names[end - 1],
                'average_score': current_avg_score,
                'window_size': window_size
            })

            if current_avg_score > best_avg_score:
                best_avg_score = current_avg_score
                best_start = start
                best_end = end
                best_window_size = window_size

            if start < len(siRNA_names) - window_size:
                current_sum = current_sum - siRNA_combined_scores[siRNA_names[start]] + siRNA_combined_scores[siRNA_names[end]]

    for window in window_data:
        window['is_best_window'] = window['start_siRNA'] == siRNA_names[best_start] and window['end_siRNA'] == siRNA_names[best_end - 1]

    # Calculate the average efficiency and safety scores for the best window
    selected_efficiency_scores = [total_scores[name] for name in siRNA_names[best_start:best_end]]
    selected_safety_scores = [norm_off_target_scores[name] for name in siRNA_names[best_start:best_end]]

    selected_dsRNA_efficiency_score = sum(selected_efficiency_scores) / len(selected_efficiency_scores)
    if safety:
        selected_dsRNA_safety_score = sum(selected_safety_scores) / len(selected_safety_scores)
    elif not safety:
        selected_dsRNA_safety_score = 0

    # Plotting data preparation
    x_positions = list(range(len(siRNA_names)))
    efficiency_scores = [total_scores[name] for name in siRNA_names]
    safety_scores = [norm_off_target_scores[name] for name in siRNA_names]
    combined_scores = [siRNA_combined_scores[name] for name in siRNA_names]

    # Plotting
    plt.figure(figsize=(12, 6))
    plt.plot(x_positions, efficiency_scores, label='Efficiency Score', linestyle='-')
    plt.plot(x_positions, safety_scores, label='Safety Score', linestyle='-')
    #plt.plot(x_positions, combined_scores, label='Combined Score', linestyle='-', color='purple')

    # Highlighting the best region
    plt.axvspan(best_start, best_end - 1, color='yellow', alpha=0.3, label='Selected Region')

    clean_name=modify_string(siRNA_names[best_start])

    # Plot settings
    plt.xlabel('Position (5 > 3)')
    plt.ylabel('Score')
    plt.title(clean_name)
    plt.legend()
    plt.grid(False)
    plt.savefig(user_dir + "/efficiency/" + clean_name + "_combined.png")

    if not os.path.exists("efficiency/region_scoring"):
        os.makedirs("efficiency/region_scoring")

    # Write window data to a text file in append mode
    #with open(user_dir + "/efficiency/region_scoring" + clean_name + '_region_scoring.txt', 'a') as file:
    #    for window in window_data:
     #       file.write(
     #           f"Start: {window['start_siRNA']}, End: {window['end_siRNA']}, Average Score: {window['average_score']}, Window Size: {window['window_size']}, Selected: {'Yes' if window['is_best_window'] else 'No'}\n")

    return {
        'start_siRNA': siRNA_names[best_start] if best_avg_score != float('-inf') else None,
        'end_siRNA': siRNA_names[best_end - 1] if best_avg_score != float('-inf') else None,
        'window_size': best_window_size,
        'selected_dsRNA_total_score': best_avg_score if best_avg_score != float('-inf') else None,
        'average_gene_total_score': gene_average,
        'selected_dsRNA_efficiency_score': selected_dsRNA_efficiency_score,
        'selected_dsRNA_safety_score': selected_dsRNA_safety_score,


    }


def modify_string(s):
    # Split the string by "_"
    parts = s.split("_")

    # Check if there are at least two "_" in the string
    if len(parts) > 2:
        # Join all parts except the last two
        new_string = "_".join(parts[:-2])
        return new_string
    else:
        # Return the original string if there are less than two "_"
        return s


def extract_subsequence(sequence, siRNA_info, bufsi=10):
    try:
        # Check if 'start_siRNA' and 'end_siRNA' keys are present
        if 'start_siRNA' not in siRNA_info or 'end_siRNA' not in siRNA_info:
            print(f"Skipping sequence due to missing 'start_siRNA' or 'end_siRNA': {sequence}")
            return None

        # Extract the start and end positions from siRNA_info
        start_match = re.search(r"_(\d+)_\d+$", siRNA_info['start_siRNA'])
        end_match = re.search(r"_\d+_(\d+)$", siRNA_info['end_siRNA'])

        if not start_match or not end_match:
            print(f"Skipping sequence due to invalid siRNA format: {sequence}")
            return None

        start_pos = int(start_match.group(1))
        end_pos = int(end_match.group(1))

        # Adjust positions based on bufsi, ensuring not to exceed sequence bounds
        if bufsi > 0:
            start_pos = max(start_pos - bufsi, 1)  # Prevent start_pos from going below 1
            end_pos = min(end_pos + bufsi, len(sequence))  # Prevent end_pos from exceeding sequence length

        # Sequence indices start at 0, but our positions start at 1.
        # Adjust slicing to include the end position correctly
        return sequence[start_pos-1:end_pos]  # end_pos is adjusted to not exceed sequence length

    except Exception as e:
        print(f"Error processing sequence: {sequence}. Error: {e}")
        return None


### bridger dsRNA stuff
def find_top_siRNA_regions(siRNA_scores, min_window, max_window, only_ORF, top_threshold):
    top_regions = []
    siRNA_indices = {}

    # Filter siRNA names based on the only_ORF flag and create an index map
    if only_ORF:
        siRNA_names = [name for name in siRNA_scores['siRNA_features'] if siRNA_scores['siRNA_features'][name]['ORF_info'] == 'ORF']
    else:
        siRNA_names = list(siRNA_scores['siRNA_features'].keys())

    if not siRNA_names:
        return None

    # Create a dictionary of siRNA names to their indices for quick lookup
    for index, name in enumerate(siRNA_names):
        siRNA_indices[name] = index

    # Calculate total scores for each siRNA
    siRNA_total_scores = {name: siRNA_scores['siRNA_features'][name]['total_score'] for name in siRNA_names}
    gene_average = sum(siRNA_total_scores.values()) / len(siRNA_names)

    for window_size in range(min_window, max_window + 1):
        current_sum = sum([siRNA_total_scores[name] for name in siRNA_names[:window_size]])

        for start in range(len(siRNA_names) - window_size + 1):
            end = start + window_size
            current_avg_score = current_sum / window_size

            # Check for overlap with existing top regions
            start_index = siRNA_indices[siRNA_names[start]]
            end_index = siRNA_indices[siRNA_names[end - 1]]
            overlaps = any(r['start_index'] <= end_index and r['end_index'] >= start_index for r in top_regions)

            if not overlaps:
                if len(top_regions) < top_threshold or current_avg_score > top_regions[-1]['average_score']:
                    if len(top_regions) >= top_threshold:
                        top_regions.pop()
                    top_regions.append({
                        'start_siRNA': siRNA_names[start],
                        'end_siRNA': siRNA_names[end - 1],
                        'average_score': current_avg_score,
                        'window_size': window_size,
                        'start_index': start_index,  # Store indices for overlap checking
                        'end_index': end_index
                    })
                    top_regions.sort(key=lambda x: x['average_score'], reverse=True)

            if start < len(siRNA_names) - window_size:
                current_sum = current_sum - siRNA_total_scores[siRNA_names[start]] + siRNA_total_scores[siRNA_names[end]]

    return top_regions


def extract_sequences_multiple(sequence, siRNA_infos, pre_gene_name, bufsi=0):
    records = []

    gene_name = pre_gene_name

    for siRNA_info in siRNA_infos:
        try:
            if 'start_siRNA' not in siRNA_info or 'end_siRNA' not in siRNA_info:
                print(f"Skipping sequence due to missing siRNA info.")
                continue

            start_match = re.search(r"_(\d+)_\d+$", siRNA_info['start_siRNA'])
            end_match = re.search(r"_\d+_(\d+)$", siRNA_info['end_siRNA'])

            if not start_match or not end_match:
                print(f"Skipping sequence due to invalid siRNA format.")
                continue

            start_pos = int(start_match.group(1))
            end_pos = int(end_match.group(1))

            if bufsi > 0:
                start_pos = max(start_pos - bufsi, 1)
                end_pos = min(end_pos + bufsi, len(sequence))

            seq_id = f"{gene_name}_{start_pos}_{end_pos}"
            extracted_sequence = sequence[start_pos-1:end_pos]
            fragment_score = siRNA_info['average_score']
            records.append({'seq_id': seq_id, 'seq': str(extracted_sequence), 'fragment_score': fragment_score})

        except Exception as e:
            print(f"Error processing sequence. Error: {e}")

    print("multiple_extracted")
    print(records)
    return records


def identify_overlaps(sequences):
    overlaps = {}
    max_overlap = 10  # Check overlaps up to 10 nucleotides

    for i in range(len(sequences)):
        seq_id1 = sequences[i]['seq_id']
        seq1 = str(sequences[i]['seq'])

        for j in range(i + 1, len(sequences)):
            seq_id2 = sequences[j]['seq_id']
            seq2 = str(sequences[j]['seq'])

            # Check for overlaps of different lengths up to 10 nt
            for overlap_length in range(1, max_overlap + 1):
                # Compare front of seq1 to back of seq2
                if seq1[:overlap_length] == seq2[-overlap_length:]:
                    overlaps.setdefault((seq_id1, seq_id2), {'type': 'front-to-back', 'length': 0, 'overlap': ''})
                    if overlaps[(seq_id1, seq_id2)]['length'] < overlap_length:
                        overlaps[(seq_id1, seq_id2)] = {
                            'type': 'front-to-back',
                            'length': overlap_length,
                            'overlap': seq1[:overlap_length]
                        }

                # Compare back of seq1 to front of seq2
                if seq1[-overlap_length:] == seq2[:overlap_length]:
                    overlaps.setdefault((seq_id1, seq_id2), {'type': 'back-to-front', 'length': 0, 'overlap': ''})
                    if overlaps[(seq_id1, seq_id2)]['length'] < overlap_length:
                        overlaps[(seq_id1, seq_id2)] = {
                            'type': 'back-to-front',
                            'length': overlap_length,
                            'overlap': seq1[-overlap_length:]
                        }

    return overlaps

def merge_sequences_with_report(sequences, overlaps):
    from collections import defaultdict

    # Convert list of sequences to a dictionary for quick access
    seq_dict = {seq['seq_id']: {'seq': str(seq['seq']), 'score': seq['fragment_score']} for seq in sequences}

    # Track merged sequences to avoid duplication
    merged = set()
    included_fragments = set()

    # Helper function to merge two sequences and track the merge details
    def merge(seq1, seq2, overlap_length, type):
        if type == 'front-to-back':
            return seq2 + seq1[overlap_length:]
        elif type == 'back-to-front':
            return seq1 + seq2[overlap_length:]

    # Initialize the final sequence list and report data
    superstring_parts = []
    merge_report = []
    total_overlap_length = 0
    detailed_merge_report = []

    # Track all current sequence IDs
    current_seq_ids = set(seq_dict.keys())

    while overlaps:
        # Sort overlaps by their length in descending order
        sorted_overlaps = sorted(overlaps.items(), key=lambda x: -x[1]['length'])
        used_seqs = set()
        merged_in_iteration = False

        # Iterate over sorted overlaps to merge sequences
        for (seq_id1, seq_id2), data in sorted_overlaps:
            if seq_id1 not in merged and seq_id2 not in merged:
                overlap_length = data['length']
                overlap_type = data['type']

                # Merge sequences
                merged_sequence = merge(seq_dict[seq_id1]['seq'], seq_dict[seq_id2]['seq'], overlap_length, overlap_type)

                # Create a new sequence ID for the merged sequence
                new_seq_id = f"{seq_id1}_{seq_id2}"
                seq_dict[new_seq_id] = {'seq': merged_sequence, 'score': (seq_dict[seq_id1]['score'] + seq_dict[seq_id2]['score']) / 2}

                # Mark original sequences as merged
                merged.add(seq_id1)
                merged.add(seq_id2)

                # Track merge details
                merge_report.append({
                    'upstream': seq_id1,
                    'downstream': seq_id2,
                    'overlap_length': overlap_length,
                    'overlap_sequence': data['overlap']
                })

                # Add to detailed merge report
                if not detailed_merge_report:
                    detailed_merge_report.append(f"{seq_id1}:{overlap_length}:{seq_id2}")
                else:
                    last_entry = detailed_merge_report.pop()
                    detailed_merge_report.append(f"{last_entry}:{overlap_length}:{seq_id2}")

                # Accumulate total overlap length
                total_overlap_length += overlap_length

                # Track included fragments
                included_fragments.add(seq_id1)
                included_fragments.add(seq_id2)
                included_fragments.add(new_seq_id)

                # Track used sequences
                used_seqs.add(seq_id1)
                used_seqs.add(seq_id2)

                # Update current sequence IDs
                current_seq_ids.add(new_seq_id)
                current_seq_ids.remove(seq_id1)
                current_seq_ids.remove(seq_id2)

                merged_in_iteration = True
                # Break after the first merge to re-evaluate overlaps
                break

        if not merged_in_iteration:
            # If no merges were made, handle the non-overlapping sequences
            remaining_seq_ids = list(current_seq_ids)
            while remaining_seq_ids:
                if not detailed_merge_report or detailed_merge_report[-1][-1].isdigit():
                    detailed_merge_report.append(f"{remaining_seq_ids.pop(0)}")
                else:
                    detailed_merge_report[-1] += f":0:{remaining_seq_ids.pop(0)}"

            break

        # Remove used overlaps
        overlaps = {(seq1, seq2): data for (seq1, seq2), data in overlaps.items() if seq1 not in used_seqs and seq2 not in used_seqs}

        # Recalculate overlaps for the newly formed sequences
        new_overlaps = {}
        for new_seq_id in current_seq_ids:
            for other_seq_id in current_seq_ids:
                if new_seq_id == other_seq_id:
                    continue
                seq1 = seq_dict[new_seq_id]['seq']
                seq2 = seq_dict[other_seq_id]['seq']
                for overlap_length in range(1, 11):  # Check overlaps up to 10 nucleotides
                    if seq1[:overlap_length] == seq2[-overlap_length:]:
                        new_overlaps[(new_seq_id, other_seq_id)] = {
                            'type': 'front-to-back',
                            'length': overlap_length,
                            'overlap': seq1[:overlap_length]
                        }
                    if seq1[-overlap_length:] == seq2[:overlap_length]:
                        new_overlaps[(new_seq_id, other_seq_id)] = {
                            'type': 'back-to-front',
                            'length': overlap_length,
                            'overlap': seq1[-overlap_length:]
                        }
        overlaps = new_overlaps

    # Add any remaining unmerged sequences to the superstring parts
    for seq_id in current_seq_ids:
        if seq_id not in merged:
            superstring_parts.append(seq_dict[seq_id]['seq'])
            included_fragments.add(seq_id)

    # Build the final superstring by concatenating all parts
    final_superstring = ''.join(superstring_parts)

    # Calculate the average score of included fragments
    included_scores = [seq_dict[seq_id]['score'] for seq_id in included_fragments if seq_id in seq_dict]
    average_score = sum(included_scores) / len(included_scores) if included_scores else 0

    # Determine excluded fragments
    all_seq_ids = {seq['seq_id'] for seq in sequences}
    excluded_fragments = all_seq_ids - included_fragments

    # Generate detailed report
    detailed_report = {
        'merged_superstring': final_superstring,
        'total_overlap_length': total_overlap_length,
        'average_fragment_score': average_score,
        'merges': merge_report,
        'included_fragments': list(included_fragments),
        'excluded_fragments': list(excluded_fragments),
        'detailed_merge_report': detailed_merge_report
    }

    return detailed_report



def bridged_report_output(sequences, path_to_output):

    overlaps = identify_overlaps(sequences)
    report = merge_sequences_with_report(sequences, overlaps)
    sequences
    with open(path_to_output + "/bridged_dsRNA.txt", "w") as report_file:
        report_file.write(">Bridger_dsRNA:\n")
        report_file.write(report['merged_superstring'] + "\n\n")
        report_file.write("Total Overlap Length:\n")
        report_file.write(str(report['total_overlap_length']) + "\n\n")
        report_file.write("Average Fragment Score:\n")
        report_file.write(str(report['average_fragment_score']) + "\n\n")

        report_file.write("Fragment details:\n")
        # Format and write the sequences
        for fragment in sequences:
            fragment_details = f"ID: {fragment['seq_id']}, Sequence: {fragment['seq']}, Score: {fragment['fragment_score']}\n"
            report_file.write(fragment_details)
        report_file.write("\n\n")

        report_file.write("\nOverlap Details:\n")
        for (seq_id1, seq_id2), data in overlaps.items():
            report_file.write(f"Overlap found between {seq_id1} and {seq_id2}: {data}\n")






#### test for dsRIP dsRNA test####

#best_region_all("Diabrotica_virgifera_lethal_genes.fasta", "Aphis_mellifera,Bombus_terrestris,Coccinella_septempunctata,Daphnia_magna,Chrysoperla_carnea")



'''''
def bridger_dsRNA_all(fasta, off_target_species, user_dir, efficiency_file,  user_fragment_length_min1=40, user_fragment_number=12, user_siRNA_length=21, only_ORF=True, ORF_corr=False):
    user_fragment_length = user_fragment_length_min1 + 1


    final_clean = {}
    all_best_dsRNA_regions = []  # List to collect all best dsRNA regions from all genes



    siRNA_length = int(user_siRNA_length)
    min_window = user_fragment_length
    max_window = user_fragment_length
    file_path = fasta

    try:
        sequence_data = read_fasta_file(file_path)
    except Exception as e:
        print(f"An error occurred: {e}")
        return

    fasta_dict = {}
    fasta_dict_score = {}

    for name, sequence in sequence_data:
        if not check_valid_dna(sequence):
            print("\nInvalid DNA sequence detected:", sequence)
            continue

        name = sanitize_sheet_title(name)

        try:
            if ORF_corr:
                # Process the sequence and find the ORF
                corrected_sequence = process_sequence(sequence, name)
                gene_ORF = ORF_finder_module.get_ORF(corrected_sequence, name)[4]
            else:
                # Use the original sequence and find the ORF
                corrected_sequence = sequence
                gene_ORF = ORF_finder_module.get_ORF(corrected_sequence, name)[4]

            # Convert DNA sequence to RNA, replacing 'T' with 'U'
            cor_seq_RNA = corrected_sequence.replace("T", "U")

        except Exception as e:
            # Handle any exceptions that may occur and return the original sequence
            print(f"Error processing sequence: {e}")
            cor_seq_RNA = sequence.replace("T", "U")  # Convert original sequence as a fallback

        return corrected_sequence, gene_ORF, cor_seq_RNA

        seq_length = len(corrected_sequence)
        ORF_info = ORF_siRNA(cor_seq_RNA)
        siRNAs, siRNAs_DNA_off_target = generate_siRNA(cor_seq_RNA, siRNA_length=siRNA_length)

        try:
            per_species_summary, all_off_targets_sum, lethal_off_targets_sum, all_siRNAs_report, off_targets_plot = off_target_siRNA_all(siRNAs_DNA_off_target, seq_length, name, off_target_species, user_dir)
            if save_off_targets_plot:
                off_targets_plot = plt.savefig(user_dir + "/efficiency/" + name + '_off_target_plot.png')
        except Exception as e:
            print(f"An error occurred during off-target analysis: {e}")
            all_siRNAs_report = []

        #access = mRNA_accessibility(cor_seq_RNA, siRNA_length=siRNA_length)
        gene_siRNA_feature = siRNA_feature_prediction(siRNAs, cor_seq_RNA, name, siRNA_length, access, ORF_info, all_siRNAs_report)
        gene_dict = {"gene_name": name, "input_gene_seq": sequence.upper(), "corrected_gene_seq": corrected_sequence, "gene_ORF": gene_ORF, "gene_RNA_seq": cor_seq_RNA, "siRNA_features": gene_siRNA_feature}
        scored_features = score_and_update_siRNA_features(gene_dict)
        best_regions = find_top_siRNA_regions(scored_features, min_window=min_window - (siRNA_length - 1), max_window=max_window - (siRNA_length - 1), only_ORF=only_ORF, top_threshold= int(user_fragment_number))
        best_dsRNA_region = extract_sequences_multiple(corrected_sequence, best_regions, pre_gene_name=name, bufsi=0)
        for entry in best_dsRNA_region:
            all_best_dsRNA_regions.append(entry)


    merge_options = bridged_report_output(all_best_dsRNA_regions, path_to_output=user_dir)  # Merge all collected best regions together


    return merge_options

'''''